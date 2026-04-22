import pyotp
import io
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.utils import timezone
from django.core.cache import cache
from django.db.models import Avg, Count, Max, Min, Q, Sum
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, mixins, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Nation, Level, Crewman, Tank, BattleRecord
from .serializers import (
    NationSerializer, LevelSerializer, CrewmanSerializer,
    TankSerializer, BattleRecordSerializer
)
from .permissions import IsStaffAnd2FAVerified

class NationViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                    mixins.CreateModelMixin, mixins.UpdateModelMixin,
                    mixins.DestroyModelMixin, GenericViewSet):
    queryset = Nation.objects.all()
    serializer_class = NationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['name']

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsStaffAnd2FAVerified]
        else:
            self.permission_classes = [permissions.AllowAny]
        return super().get_permissions()

    @action(detail=False, methods=['GET'], url_path='stats')
    def nation_stats(self, request):
        stats = Nation.objects.annotate(
            tanks_count=Count('tanks')
        ).aggregate(
            total_nations=Count('id'),
            total_tanks=Sum('tanks_count'),
            avg_tanks_per_nation=Avg('tanks_count')
        )
        
        most_popular = Nation.objects.annotate(
            tanks_count=Count('tanks')
        ).order_by('-tanks_count').first()
        
        stats['most_popular_nation'] = most_popular.name if most_popular else None
        stats['most_popular_tanks_count'] = most_popular.tanks_count if most_popular else 0
        
        if stats['avg_tanks_per_nation'] is None:
            stats['avg_tanks_per_nation'] = 0
        
        return Response(stats)

class LevelViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin,
                   mixins.CreateModelMixin, mixins.UpdateModelMixin,
                   mixins.DestroyModelMixin, GenericViewSet):
    queryset = Level.objects.all()
    serializer_class = LevelSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['level_number']

    def create(self, request, *args, **kwargs):
        print("=== CREATE LEVEL ===")
        print("Request data:", request.data)
        return super().create(request, *args, **kwargs)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsStaffAnd2FAVerified]
        else:
            self.permission_classes = [permissions.AllowAny]
        return super().get_permissions()

    @action(detail=False, methods=['GET'], url_path='stats')
    def level_stats(self, request):
        stats = Level.objects.aggregate(
            total_levels=Count('id'),
            min_level=Min('level_number'),
            max_level=Max('level_number'),
            avg_level=Avg('level_number')
        )
        
        level_distribution = Level.objects.annotate(
            tanks_count=Count('tanks')
        ).values('level_number', 'tanks_count').order_by('level_number')
        
        stats['level_distribution'] = list(level_distribution)
        
        if stats['avg_level'] is None:
            stats['avg_level'] = 0
        
        return Response(stats)
    
    def destroy(self, request, *args, **kwargs):
        level = self.get_object()
        
        max_level = Level.objects.aggregate(Max('level_number'))['level_number__max']
        
        if level.level_number != max_level:
            return Response(
                {'error': f'Нельзя удалить уровень {level.level_number}. Можно удалить только последний уровень ({max_level}).'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        tanks = Tank.objects.filter(level=level)
        
        total_refund = sum(tank.level.creation_cost for tank in tanks)
        
        for tank in tanks:
            if tank.is_in_battle:
                battle = BattleRecord.objects.filter(tank=tank, result='pending').first()
                if battle:
                    battle.result = 'defeat'
                    battle.reward_earned = 0
                    battle.finished_at = timezone.now()
                    battle.save()
                tank.is_in_battle = False
                tank.save()  
            
            tank.owner.credits += tank.level.creation_cost
            tank.owner.save()
            tank.delete()
        
        level.delete()
        
        return Response({'success': True, 'message': f'Уровень {level.level_number} удалён. Возвращено кредитов: {total_refund}'})

class CrewmanViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = CrewmanSerializer
    queryset = Crewman.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['user__username', 'credits', 'garage_slots']

    def get_permissions(self):
        if self.action in ['update', 'partial_update', 'list_users']:
            return [IsStaffAnd2FAVerified()]
        if self.action == 'register':
            return [permissions.AllowAny()]  
        return super().get_permissions()

    def get_object(self):
        crewman, created = Crewman.objects.get_or_create(user=self.request.user)
        return crewman

    @action(detail=False, methods=['GET'], url_path='profile')
    def profile(self, request):
        crewman = self.get_object()
        serializer = self.get_serializer(crewman)
        return Response(serializer.data)

    @action(detail=False, methods=['POST'], url_path='expand-garage')
    def expand_garage(self, request):
        crewman = self.get_object()
        cost = 500
        if crewman.credits >= cost:
            crewman.credits -= cost
            crewman.garage_slots += 1
            crewman.save()
            return Response({'success': True, 'garage_slots': crewman.garage_slots})
        return Response({'success': False, 'error': 'Not enough credits'}, status=status.HTTP_400_BAD_REQUEST)

    @method_decorator(csrf_exempt)
    @action(detail=False, methods=['POST'], url_path='register')
    def register(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        email = request.data.get('email', '')
        
        print(f"=== REGISTER ATTEMPT: {username} ===")  # для отладки
        
        if not username or not password:
            return Response({'error': 'Username and password required'}, status=status.HTTP_400_BAD_REQUEST)
        
        if User.objects.filter(username=username).exists():
            return Response({'error': 'Username already exists'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = User.objects.create_user(username=username, password=password, email=email)
            crewman = Crewman.objects.create(user=user, credits=1000, garage_slots=3)
            login(request, user)
            serializer = self.get_serializer(crewman)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        except Exception as e:
            print(f"Registration error: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['GET'], url_path='list-users')
    def list_users(self, request):
        crewmen = Crewman.objects.select_related('user').all()
        data = [{
            'id': c.id,
            'username': c.user.username,
            'credits': c.credits,
            'garage_slots': c.garage_slots,
            'tanks_count': c.tanks.count()
        } for c in crewmen]
        return Response(data)

    def update(self, request, pk=None):
        try:
            crewman = Crewman.objects.get(pk=pk)
        except Crewman.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)
        
        credits = request.data.get('credits')
        garage_slots = request.data.get('garage_slots')
        
        if credits is not None:
            crewman.credits = credits
        if garage_slots is not None:
            crewman.garage_slots = garage_slots
        
        crewman.save()
        return Response({
            'success': True, 
            'credits': crewman.credits, 
            'garage_slots': crewman.garage_slots
        })

class TankViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, mixins.UpdateModelMixin,
                  mixins.DestroyModelMixin, GenericViewSet):
    serializer_class = TankSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['owner__user__username', 'owner__id', 'nation__name', 'level__level_number', 'is_in_battle']

    def get_queryset(self):
        qs = Tank.objects.select_related('owner__user', 'nation', 'level')
        user = self.request.user
        if not user.is_superuser or not cache.get(f'2fa_{user.id}', False):
            crewman, _ = Crewman.objects.get_or_create(user=user)
            return qs.filter(owner=crewman)
        return qs.all()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def update(self, request, *args, **kwargs):
        tank = self.get_object()
        
        if tank.owner.user != request.user and not request.user.is_superuser:
            return Response({'error': 'Permission denied'}, status=status.HTTP_403_FORBIDDEN)
        
        if request.user.is_superuser and tank.owner.user != request.user:
            if not cache.get(f'2fa_{request.user.id}', False):
                return Response({'error': '2FA required for editing other users tanks'}, 
                                status=status.HTTP_403_FORBIDDEN)
        
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        return self.update(request, *args, **kwargs)

    @action(detail=False, methods=['POST'], url_path='create')
    def create_tank(self, request):
        name = request.data.get('name')
        nation_id = request.data.get('nation')
        level_id = request.data.get('level')
        picture = request.FILES.get('picture')
        
        if not all([name, nation_id, level_id]):
            return Response({'error': 'Missing fields'}, status=status.HTTP_400_BAD_REQUEST)
        
        crewman, created = Crewman.objects.get_or_create(user=self.request.user)
        if crewman.available_garage_slots() <= 0:
            return Response({'error': 'No free garage slots'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            nation = Nation.objects.get(id=nation_id)
            level = Level.objects.get(id=level_id)
        except (Nation.DoesNotExist, Level.DoesNotExist):
            return Response({'error': 'Invalid nation or level'}, status=status.HTTP_400_BAD_REQUEST)
        
        cost = level.creation_cost
        if crewman.credits < cost:
            return Response({'error': f'Not enough credits. Need {cost}, you have {crewman.credits}'},
                            status=status.HTTP_400_BAD_REQUEST)
        
        tank = Tank.objects.create(
            name=name,
            owner=crewman,
            nation=nation,
            level=level,
            is_in_battle=False,
            picture=picture
        )
        crewman.credits -= cost
        crewman.save()
        serializer = self.get_serializer(tank)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['POST'], url_path='upgrade')
    def upgrade_tank(self, request, pk=None):
        tank = self.get_object()
        print(f"=== UPGRADE ATTEMPT ===")
        print(f"Tank: {tank.name}, Level: {tank.level.level_number}")
        print(f"Is in battle: {tank.is_in_battle}")
        print(f"Can upgrade: {tank.can_upgrade()}")
        
        if tank.is_in_battle:
            return Response({'error': 'Танк в бою'}, status=status.HTTP_400_BAD_REQUEST)
        
        if not tank.can_upgrade():
            return Response({'error': 'Нельзя улучшить (макс. уровень или танк в бою)'}, status=status.HTTP_400_BAD_REQUEST)
        
        next_level = Level.objects.filter(level_number=tank.level.level_number + 1).first()
        print(f"Next level: {next_level}")
        
        if not next_level:
            return Response({'error': 'Следующий уровень не найден'}, status=status.HTTP_400_BAD_REQUEST)
        
        cost = next_level.creation_cost//2
        print(f"Upgrade cost: {cost}")
        print(f"User credits: {tank.owner.credits}")
        
        if cost is None:
            return Response({'error': 'Стоимость улучшения не определена'}, status=status.HTTP_400_BAD_REQUEST)
        
        if tank.owner.credits < cost:
            return Response({'error': f'Не хватает кредитов. Нужно {cost}, у вас {tank.owner.credits}'}, 
                            status=status.HTTP_400_BAD_REQUEST)
        
        tank.level = next_level
        tank.owner.credits -= cost
        tank.owner.save()
        tank.save()
        
        return Response({'success': True, 'new_level': tank.level.level_number})

    @action(detail=True, methods=['POST'], url_path='sell')
    def sell_tank(self, request, pk=None):
        tank = self.get_object()
        if tank.is_in_battle:
            return Response({'error': 'Tank is in battle'}, status=status.HTTP_400_BAD_REQUEST)
        price = tank.sell_price
        tank.sell()
        return Response({'success': True, 'credits_earned': price})

    @action(detail=False, methods=['GET'], url_path='stats')
    def tank_stats(self, request):
        stats = Tank.objects.aggregate(
            total_tanks=Count('id'),
            avg_level=Avg('level__level_number'),
            max_level=Max('level__level_number'),
            min_level=Min('level__level_number'),
        )
        nation_popular = Tank.objects.values('nation__name').annotate(cnt=Count('id')).order_by('-cnt').first()
        stats['most_popular_nation'] = nation_popular['nation__name'] if nation_popular else None
        
        if stats['avg_level'] is None:
            stats['avg_level'] = 0
        
        return Response(stats)

    @action(detail=False, methods=['GET'], url_path='export-excel')
    def export_tanks_excel(self, request):
        tanks = self.get_queryset()
        
        owner_id = request.GET.get('owner')
        if owner_id:
            tanks = tanks.filter(owner_id=owner_id)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Танки"
        headers = ['ID', 'Название', 'Владелец', 'Нация', 'Уровень', 'Урон/мин', 'В бою', 'Создан']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row, tank in enumerate(tanks, 2):
            ws.cell(row=row, column=1, value=tank.id)
            ws.cell(row=row, column=2, value=tank.name)
            ws.cell(row=row, column=3, value=tank.owner.user.username)
            ws.cell(row=row, column=4, value=tank.nation.name)
            ws.cell(row=row, column=5, value=tank.level.level_number)
            ws.cell(row=row, column=6, value=tank.dpm)
            ws.cell(row=row, column=7, value='Да' if tank.is_in_battle else 'Нет')
            ws.cell(row=row, column=8, value=tank.created_at.strftime('%Y-%m-%d %H:%M'))
        
        for column in ws.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="tanks.xlsx"'
        return response

class BattleRecordViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, GenericViewSet):
    serializer_class = BattleRecordSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['result', 'battle_level__level_number', 'tank__name']

    def get_queryset(self):
        crewman, created = Crewman.objects.get_or_create(user=self.request.user)
        if self.request.user.is_superuser:
            return BattleRecord.objects.all().order_by('-started_at')
        return BattleRecord.objects.filter(crewman=crewman).order_by('-started_at')

    @action(detail=False, methods=['POST'], url_path='start')
    def start_battle(self, request):
        tank_id = request.data.get('tank_id')
        battle_level_id = request.data.get('battle_level_id')
        
        if not tank_id or not battle_level_id:
            return Response({'error': 'tank_id and battle_level_id required'}, status=status.HTTP_400_BAD_REQUEST)
        
        crewman, _ = Crewman.objects.get_or_create(user=self.request.user)
        
        try:
            tank = Tank.objects.get(id=tank_id, owner=crewman)
            battle_level = Level.objects.get(id=battle_level_id)
        except (Tank.DoesNotExist, Level.DoesNotExist):
            return Response({'error': 'Invalid tank or level'}, status=status.HTTP_400_BAD_REQUEST)
        
        if tank.is_in_battle:
            return Response({'error': 'Tank already in battle'}, status=status.HTTP_400_BAD_REQUEST)
        
        tank_lvl = tank.level.level_number
        if battle_level.level_number not in [tank_lvl - 1, tank_lvl, tank_lvl + 1]:
            return Response({'error': 'Battle level not available for this tank'}, status=status.HTTP_400_BAD_REQUEST)
        
        battle_record = BattleRecord.objects.create(
            tank=tank,
            crewman=crewman,
            battle_level=battle_level,
            result='pending',
            reward_earned=0
        )
        tank.is_in_battle = True
        tank.save()
        
        serializer = BattleRecordSerializer(battle_record)
        return Response({
            'id': battle_record.id,
            'battle_duration': tank_lvl,  
            **serializer.data
        }, status=status.HTTP_201_CREATED)
        
    @action(detail=True, methods=['GET'], url_path='remaining-time')
    def remaining_time(self, request, pk=None):
        battle = self.get_object()
        if battle.result != 'pending':
            return Response({'remaining': 0, 'finished': True, 'result': battle.result})
        
        from django.utils import timezone
        elapsed = (timezone.now() - battle.started_at).total_seconds()
        tank_level = battle.tank.level.level_number
        total_duration = tank_level
        remaining = max(0, total_duration - elapsed)
        
        return Response({
            'remaining': int(remaining),
            'finished': False,
            'total_duration': total_duration
        })
        
    @action(detail=True, methods=['POST'], url_path='claim')
    def claim_battle(self, request, pk=None):
        battle = self.get_object()
        
        if battle.result != 'pending':
            return Response({'error': 'Бой уже завершён'}, status=status.HTTP_400_BAD_REQUEST)
        
        from django.utils import timezone
        elapsed = (timezone.now() - battle.started_at).total_seconds()
        tank_level = battle.tank.level.level_number
        required_duration = tank_level  # секунд
        
        if elapsed < required_duration:
            remaining = int(required_duration - elapsed)
            return Response({
                'error': f'Бой ещё не закончился. Осталось {remaining} секунд.',
                'remaining': remaining
            }, status=status.HTTP_400_BAD_REQUEST)
        
        tank_lvl = battle.tank.level.level_number
        battle_lvl = battle.battle_level.level_number
        
        if battle_lvl < tank_lvl:
            win_chance = 100
        elif battle_lvl == tank_lvl:
            win_chance = 75
        else:
            win_chance = 25
        
        import random
        is_victory = random.randint(1, 100) <= win_chance

        battle.process_battle_result(is_victory)
        
        serializer = BattleRecordSerializer(battle)
        return Response(serializer.data)
    
    @action(detail=True, methods=['GET'], url_path='result')
    def battle_result(self, request, pk=None):
        battle = self.get_object()
        serializer = BattleRecordSerializer(battle)
        return Response(serializer.data)

    @action(detail=False, methods=['GET'], url_path='stats')
    def battle_stats(self, request):
        crewman, created = Crewman.objects.get_or_create(user=self.request.user)
        if self.request.user.is_superuser:
            queryset = BattleRecord.objects.all()
        else:
            queryset = BattleRecord.objects.filter(crewman=crewman)
        
        stats = queryset.aggregate(
            total_battles=Count('id'),
            victories=Count('id', filter=Q(result='victory')),
            defeats=Count('id', filter=Q(result='defeat')),
            total_reward=Sum('reward_earned'),
            avg_reward=Avg('reward_earned')
        )
        stats['win_rate'] = round(stats['victories'] / stats['total_battles'] * 100, 1) if stats['total_battles'] else 0
        
        if stats['avg_reward'] is None:
            stats['avg_reward'] = 0
        
        return Response(stats)
    
    @action(detail=True, methods=['POST'], url_path='cancel')
    def cancel_battle(self, request, pk=None):
        battle = self.get_object()
        
        if battle.result != 'pending':
            return Response({'error': 'Бой уже завершён'}, status=status.HTTP_400_BAD_REQUEST)
        
        battle.result = 'defeat'
        battle.reward_earned = 0
        battle.finished_at = timezone.now()
        battle.save()
        
        battle.tank.is_in_battle = False
        battle.tank.save()
        
        return Response({'success': True, 'message': 'Бой отменён'})

class UserViewSet(viewsets.GenericViewSet):
    permission_classes = [permissions.AllowAny]

    @action(detail=False, methods=['GET'], url_path='info')
    def get_info(self, request):
        data = {
            'username': request.user.username if request.user.is_authenticated else None,
            'is_authenticated': request.user.is_authenticated,
            'is_staff': request.user.is_staff,
            'crewman_id': None,   
        }
        if request.user.is_authenticated:
            crewman, created = Crewman.objects.get_or_create(user=request.user)
            data['credits'] = crewman.credits
            data['garage_slots'] = crewman.garage_slots
            data['second'] = cache.get(f'2fa_{request.user.id}', False)
            data['crewman_id'] = crewman.id
        return Response(data)

    @method_decorator(csrf_exempt)
    @action(detail=False, methods=['POST'], url_path='login')
    def login_user(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        print(f"=== LOGIN ATTEMPT: {username} ===")  
        
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            print(f"Login successful: {user.username}")
            return Response({'success': True})
        
        print(f"Login failed for: {username}")
        return Response({'success': False}, status=status.HTTP_401_UNAUTHORIZED)

    @action(detail=False, methods=['POST'], url_path='logout')
    def logout_user(self, request):
        logout(request)
        return Response({'success': True})

    @action(detail=False, methods=['GET'], url_path='get-totp')
    def get_totp(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Not authenticated'}, status=401)
        crewman = Crewman.objects.get(user=request.user)
        if not crewman.totp_key:
            crewman.totp_key = pyotp.random_base32()
            crewman.save()
        totp = pyotp.TOTP(crewman.totp_key)
        provisioning_uri = totp.provisioning_uri(name=request.user.username, issuer_name="TankBattles")
        return Response({'url': provisioning_uri})

    @action(detail=False, methods=['POST'], url_path='second-login')
    def second_login(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Not authenticated'}, status=401)
        key = request.data.get('key')
        if not key:
            return Response({'error': 'Key required'}, status=400)
        crewman = Crewman.objects.get(user=request.user)
        if not crewman.totp_key:
            return Response({'error': 'TOTP not set up'}, status=400)
        totp = pyotp.TOTP(crewman.totp_key)
        if totp.verify(key):
            cache.set(f'2fa_{request.user.id}', True, timeout=600)  
            return Response({'success': True})
        return Response({'success': False}, status=400)

    @action(detail=False, methods=['POST'], url_path='logout-2fa')
    def logout_2fa(self, request):
        if request.user.is_authenticated:
            cache.delete(f'2fa_{request.user.id}')
        return Response({'success': True})
    
    @action(detail=False, methods=['GET'], url_path='export-all-stats')
    def export_all_stats(self, request):
        wb = Workbook()
        
        ws_tanks = wb.active
        ws_tanks.title = "Танки"
        tanks = Tank.objects.select_related('owner__user', 'nation', 'level').all()
        headers_tanks = ['ID', 'Название', 'Владелец', 'Нация', 'Уровень', 'Урон/мин', 'В бою', 'Создан']
        for col, header in enumerate(headers_tanks, 1):
            cell = ws_tanks.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row, tank in enumerate(tanks, 2):
            ws_tanks.cell(row=row, column=1, value=tank.id)
            ws_tanks.cell(row=row, column=2, value=tank.name)
            ws_tanks.cell(row=row, column=3, value=tank.owner.user.username)
            ws_tanks.cell(row=row, column=4, value=tank.nation.name)
            ws_tanks.cell(row=row, column=5, value=tank.level.level_number)
            ws_tanks.cell(row=row, column=6, value=tank.dpm)
            ws_tanks.cell(row=row, column=7, value='Да' if tank.is_in_battle else 'Нет')
            ws_tanks.cell(row=row, column=8, value=tank.created_at.strftime('%Y-%m-%d %H:%M'))
        
        ws_battles = wb.create_sheet("Бои")
        battles = BattleRecord.objects.select_related('tank', 'crewman__user', 'battle_level').all()
        headers_battles = ['ID', 'Танк', 'Игрок', 'Уровень боя', 'Результат', 'Награда', 'Начало', 'Завершение']
        for col, header in enumerate(headers_battles, 1):
            cell = ws_battles.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row, battle in enumerate(battles, 2):
            ws_battles.cell(row=row, column=1, value=battle.id)
            ws_battles.cell(row=row, column=2, value=battle.tank.name)
            ws_battles.cell(row=row, column=3, value=battle.crewman.user.username)
            ws_battles.cell(row=row, column=4, value=battle.battle_level.level_number)
            ws_battles.cell(row=row, column=5, value='Победа' if battle.result == 'victory' else ('Поражение' if battle.result == 'defeat' else 'Ожидание'))
            ws_battles.cell(row=row, column=6, value=battle.reward_earned)
            ws_battles.cell(row=row, column=7, value=battle.started_at.strftime('%Y-%m-%d %H:%M'))
            ws_battles.cell(row=row, column=8, value=battle.finished_at.strftime('%Y-%m-%d %H:%M') if battle.finished_at else '-')
        
        ws_players = wb.create_sheet("Игроки")
        crewmen = Crewman.objects.select_related('user').all()
        headers_players = ['ID', 'Имя', 'Кредиты', 'Слоты', 'Танков', 'Дата регистрации']
        for col, header in enumerate(headers_players, 1):
            cell = ws_players.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row, crew in enumerate(crewmen, 2):
            ws_players.cell(row=row, column=1, value=crew.id)
            ws_players.cell(row=row, column=2, value=crew.user.username)
            ws_players.cell(row=row, column=3, value=crew.credits)
            ws_players.cell(row=row, column=4, value=crew.garage_slots)
            ws_players.cell(row=row, column=5, value=crew.tanks.count())
            ws_players.cell(row=row, column=6, value=crew.created_at.strftime('%Y-%m-%d %H:%M'))
        
        ws_nations = wb.create_sheet("Нации")
        nations = Nation.objects.all()
        headers_nations = ['ID', 'Название', 'Количество танков']
        for col, header in enumerate(headers_nations, 1):
            cell = ws_nations.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row, nation in enumerate(nations, 2):
            ws_nations.cell(row=row, column=1, value=nation.id)
            ws_nations.cell(row=row, column=2, value=nation.name)
            ws_nations.cell(row=row, column=3, value=nation.tanks.count())
        
        ws_levels = wb.create_sheet("Уровни")
        levels = Level.objects.all()
        headers_levels = ['ID', 'Уровень', 'Стоимость создания', 'Награда за бой']
        for col, header in enumerate(headers_levels, 1):
            cell = ws_levels.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        for row, level in enumerate(levels, 2):
            ws_levels.cell(row=row, column=1, value=level.id)
            ws_levels.cell(row=row, column=2, value=level.level_number)
            ws_levels.cell(row=row, column=3, value=level.creation_cost)
            ws_levels.cell(row=row, column=4, value=level.battle_reward)
        
        for ws in [ws_tanks, ws_battles, ws_players, ws_nations, ws_levels]:
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="full_statistics.xlsx"'
        return response